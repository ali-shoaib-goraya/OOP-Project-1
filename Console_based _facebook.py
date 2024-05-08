
from openpyxl import Workbook, load_workbook
from datetime import datetime as dt
from numpy import NaN
import numpy as np
#wb = Workbook()         #Use this command to create workbook
#wb.save('Facebook.xlsx')   #Use this to save file O
wb = load_workbook('Facebook.xlsx')
wb.save('Facebook.xlsx')
class Facebook:
    wb = load_workbook('Facebook.xlsx')
    sh = wb['Sheet4']
    sh['A1'] = 'Friends_Posts'
    sh['A2'] = 'Emails'
    sh['B2'] = 'Posts'
    wb.save('Facebook.xlsx')

    wb = load_workbook('Facebook.xlsx')
    Sh = wb['Sheet']
    wb.save('Facebook.xlsx')
    Sh['A1'] = 'Email'
    Sh['B1'] = 'Password'
    Sh['C1'] = 'Name'
    Sh['D1'] = 'DOB'
    Sh['E1'] = 'Work_experience'
    Sh['F1'] = 'Bio'
    Sh['G1'] = 'Education'

    def __init__(self):
        self.name = None
        self.password = None
        self.email = None
        self.bio = None
        self.work_experience = None
        self.education = None
        self.Dob = None
        self.Logged_in = False

    def check_signup(self, email):
        self.emails= []
        end_row = Facebook.Sh.max_row+1
        for i in range(2,end_row):
            valu = Facebook.Sh.cell(i,1).value
            self.emails.append(valu)
        if email in self.emails:
            return True
        else:
            return False

################ To create account of that Roll_no which don't have account  ##################
    def Signup(self): 
        email = input('Enter your email:')
        __password= input('Enter your password:')
        name = input('Enter your name:')
        Dob = input('Enter your date of birth:')
        sign_up = Facebook.check_signup(self, email)
        if sign_up== True:
            print('You already have an account! move to login')
        else:
            self.email = email
            self.__password =__password
            self.Dob = Dob
            self.name = name
            lst = []
            lst.append(self.email)
            lst.append(self.__password)
            lst.append(self.name)
            lst.append(self.Dob)
            last_row = Facebook.Sh.max_row+1
            #last_column = Sheet.max_column+1
            for i in range(len(lst)):
                Facebook.Sh.cell(last_row,column=i+1, value=lst[i])   # To add the value in cell
            Facebook.wb.save("Facebook.xlsx")
            #sh1= Facebook.wb['Sheet_1']
            Sh1 = Facebook.wb['Sheet1']
            Sh1['A1'] = 'Friend_requests_file'
            Facebook.wb.save('Facebook.xlsx')
            last_col = Sh1.max_column+1
            Sh1.cell(1,last_col,value=self.email)
            Facebook.wb.save('Facebook.xlsx')

            Sh2 = Facebook.wb['Sheet2']
            Sh2['A1'] = 'Friend_file'
            Facebook.wb.save('Facebook.xlsx')
            last_col = Sh2.max_column+1
            Sh2.cell(1,last_col,value=self.email)
            Facebook.wb.save('Facebook.xlsx')


            sh1 = Facebook.wb['Sheet5']
            sh1['A1'] = 'Notification_file'
            last = sh1.max_column+1
            sh1.cell(1,last,value=self.email)
            Facebook.wb.save('Facebook.xlsx')

            print('Your Account is created successfully')


    def Login(self):
        email = input('Enter your email:')
        name = input('Enter your name:')
        password= input('Enter your password:')
        if self.Logged_in == False:
            sign_up = Facebook.check_signup(self,email)
            if sign_up == True:
                self.name_lst = []

                end_row = Facebook.Sh.max_row+1
                for i in range(1,end_row):
                    valu = Facebook.Sh.cell(i,3).value
                    self.name_lst.append(valu)
                if name in self.name_lst:
                    index = self.name_lst.index(name)
                    if name == self.name_lst[index]:
                        self.email_lst = []
                        end_row = Facebook.Sh.max_row+1
                        for i in range(1,end_row):
                            valu = Facebook.Sh.cell(i,1).value
                            self.email_lst.append(valu)
                        if email in self.email_lst:
                            index = self.email_lst.index(email)
                        if email == self.email_lst[index]:
                            self.password_lst = []
                            end_row = Facebook.Sh.max_row+1
                            for i in range(1,end_row):
                                p = Facebook.Sh.cell(i,2).value
                                self.password_lst.append(p) 
                            if password == self.password_lst[index]:
                                self.Logged_in = True
                                self.email = email
                                self.password = password
                                self.name = name
                                #Facebook.Sheet.cell(index+1,7, value=True)
                                #Facebook.wb.save('Facebook.xlsx')
                                #messagebox.showinfo("login","successfully logged in")   
                                print('You have successfuly logged in')           
                            else:
                                print('Invalid Password!   Try again')   
                        else: 
                            print('Invalid Email!   Try again')         
                else:
                    print('Invalid Name!   Try again')                       
            else:
                 print("You don't have an account! move to sign_up")                          
        else:
            print('You are already logged in')


    def Return_column_values_lst(self,sh,s_row,e_row,col):
        lst = []
        for row in range(s_row,e_row):
            valu = sh.cell(row,col).value
            lst.append(valu)
        return(lst)

    def Return_Row_values_lst(self,sh,s_col,e_col,row):
        lst = []
        for col in range(s_col,e_col):
            valu = sh.cell(row,col).value
            lst.append(valu)
        return(lst)

    def Return_index_in_lst(self,lst,key):
        if key in lst:
            index = lst.index(key)
            return(index)



class User:

    def __init__(self):
        self.obj1 = Facebook()


    def Add_bio(self):
        email = obj1.email
            #self.Logged_in = Facebook.check_logged_in(self,name)
        if obj1.Logged_in == True: 
            obj1.bio = input('Enter your bio:')
            email_lst = []
            end_row = Facebook.Sh.max_row+1
            for i in range(1,end_row):
                valu = Facebook.Sh.cell(i,1).value
                email_lst.append(valu)
            if email in email_lst:
                index = email_lst.index(email)
                Facebook.Sh.cell(index+1, 6, value=obj1.bio)
                Facebook.wb.save('Facebook.xlsx')
                print('Your bio has been has been added')
        else:
            print('you are not  logged_in')

    def Add_Work(self):
        email = obj1.email
            #self.Logged_in = Facebook.check_logged_in(self,name)
        if obj1.Logged_in == True: 
            obj1.work_experience = input('Enter your work experience:')
            email_lst = []
            end_row = Facebook.Sh.max_row+1
            for i in range(1,end_row):
                valu = Facebook.Sh.cell(i,1).value
                email_lst.append(valu)
            if email in email_lst:
                index = email_lst.index(email)
                Facebook.Sh.cell(index+1, 5, value=obj1.work_experience)
                Facebook.wb.save('Facebook.xlsx')
                print('Your work details has been added')
        else:
            print('you are not  logged_in')
        
    def Add_Education(self):
        email = obj1.email
            #self.Logged_in = Facebook.check_logged_in(self,name)
        if obj1.Logged_in == True: 
            obj1.education = input('Enter your Education details:')
            email_lst = []
            end_row = Facebook.Sh.max_row+1
            for i in range(1,end_row):
                valu = Facebook.Sh.cell(i,1).value
                email_lst.append(valu)
            if email in email_lst:
                index = email_lst.index(email)
                Facebook.Sh.cell(index+1, 7, value=obj1.education)
                Facebook.wb.save('Facebook.xlsx')
                print('Your education details has been added')
        else:
            print('you are not  logged_in')
    

    
    def Search_member_by_name(self):
        if obj1.Logged_in == True:
            name = input('Enter name to search profile:')
            wb = load_workbook("Facebook.xlsx")
            sh = wb['Sheet']
            names_lst = []
            last_row = sh.max_row+1
            for i in range(2,last_row):
                valu = sh.cell(i,3).value
                names_lst.append(valu)
            if name in names_lst:
                index = names_lst.index(name)
                row = index+2
                print(str(name)+"'s  profile:")
                nn = sh.cell(row, 3).value
                print('Name:', nn)
                dob = sh.cell(row,4).value
                print('Date of Birth:', dob)
                bio = sh.cell(row, 6).value
                print('Bio:', bio)
                edu = sh.cell(row, 7).value
                print('Education:', edu)
                work = sh.cell(row, 5).value
                print('Work Experience:', work)
            else:
                print('There is no member of this name!')
        else:
            print('You are not logged in!')


class Friends:

    def __init__(self):
        self.obj1 = Facebook()          ##   Compoosition  ##
        self.sent = False
        self.accept = False
        self.reject = False


    def get_max_row_in_col(self,ws, column):
        return max([cell[0] for cell in ws._cells if cell[1] == column])
    
    def Return_friends_lst(self):
        if obj1.Logged_in == True:
            user = obj1.email
            wb = load_workbook('Facebook.xlsx')
            sh = wb['Sheet2']
            last_col = sh.max_column+1
            lst = []
            for i in range(2,last_col):
                valu = sh.cell(1,i).value
                lst.append(valu)
            if user in lst:
                index = lst.index(user)
                col = index+2
                last_row = obj3.get_max_row_in_col(sh,col)+1
                friends = []
                for i in range(1,last_row):
                    valu = sh.cell(i,col).value
                    friends.append(valu)  
                return(friends)
            else:
                print(user, 'has no friends')
        else:
            print('You are not logged in!')

    def Return_Friend_requests_lst(self):
        user = obj1.email
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        last_col = sh.max_column+1
        lst = []
        for i in range(2,last_col):
            valu = sh.cell(1,i).value
            lst.append(valu)
        if user in lst:
            index = lst.index(user)
        col = index+2
        last_row = Friends.get_max_row_in_col(self,sh,col)+1
        f_lst = []
        for row in range(2,last_row):
            valu = sh.cell(row,col).value
            f_lst.append(valu)
        return(f_lst)


    def Return_Reciever_requests(self,reciever):
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        last_col = sh.max_column+1
        lst = []
        for i in range(2,last_col):
            valu = sh.cell(1,i).value
            lst.append(valu)
        if reciever in lst:
            index = lst.index(reciever)
        col = index+2
        last_row = Friends.get_max_row_in_col(self,sh,col)+1
        f_lst = []
        for row in range(2,last_row):
            valu = sh.cell(row,col).value
            f_lst.append(valu)
        return(f_lst)

    def Send_friend_request(self):
        #print(obj1.Logged_in)
        user = obj1.email
        if obj1.Logged_in == True:    ## Use of composition ##
            nn = obj1.email
            choice = input('Enter the email to whom U want to send friend request:')
            wb = load_workbook('Facebook.xlsx')
            sh1 = wb['Sheet']
            end = sh1.max_row+1
            Account_lst = obj1.Return_column_values_lst(sh1,1,end,1)
            if choice in Account_lst:
                friend_requests = Friends.Return_Friend_requests_lst(self)
                R_lst = Friends.Return_Reciever_requests(self,choice)
                if user not in R_lst:
                    if choice not in friend_requests:
                        frind_lst = Friends.Return_friends_lst(self)
                        if choice not in frind_lst:
                            wb = load_workbook('Facebook.xlsx')
                            sh = wb['Sheet1']
                            col = sh.max_column+1
                            name_lst = []
                            for i in range(2,col):
                                val = sh.cell(1,i).value
                                name_lst.append(val)
                            if choice in name_lst:
                                index = name_lst.index(choice)
                                col = index+2
                                last_row = Friends.get_max_row_in_col(self,sh,col)
                                last_row+= 1
                                sh.cell(last_row,col,value=nn)
                                wb.save('Facebook.xlsx')
                                self.sent = True
                                print('Friend request has been sent')
                                Noti = user+' has sent you a friend request'
                                sh2 = wb['Sheet5']
                                l_col = sh2.max_column+1
                                lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                                for i in range(len(lst)):
                                    if choice in lst:
                                        index = lst.index(choice)+2
                                        last_row = obj3.get_max_row_in_col(sh2,index)+1
                                        sh2.cell(last_row,index,value=Noti)
                                        wb.save('Facebook.xlsx')
                                        break
                                    else:
                                        print('user is not in header')
                            else:
                                print('There is no such member!')
                        else:
                            print('You are already friends!')
                    else:
                        print('You already have friend request this person')
                else:
                    print('You already have sent friend request!')
            else:
                print('There is no such member!')
        else:
            print('You are not logged in............')
    
    def Accept_friend_request(self):
        if obj1.Logged_in == True:
            choice = input('Enter email of member of which friend request you want to accept:')
            user  = obj1.email
            wb = load_workbook('Facebook.xlsx')
            sh = wb['Sheet1']
            col1 = sh.max_column+1
            name_lst = []
            for i in range(2,col1):
                val = sh.cell(1,i).value
                name_lst.append(val)
            if user in name_lst:
                index = name_lst.index(user)
                col = index+2
                last_row = Friends.get_max_row_in_col(self,sh,col)
                last_row+= 1
                lst = []
                for i in range(2,last_row):
                    valu = sh.cell(i,col).value
                    lst.append(valu)
                if choice in lst:
                    index = lst.index(choice)
                    row = index+2
                    sh.cell(row, col, value=NaN)
                    wb = load_workbook('Facebook.xlsx')
                    sh = wb['Sheet2']
                    col1 = sh.max_column+1
                    name_lst = []
                    for i in range(2,col1):
                        val = sh.cell(1,i).value
                        name_lst.append(val)
                    index = name_lst.index(user)
                    index2 = name_lst.index(choice)
                    col = index+2
                    col2 = index2+2
                    last_row = Friends.get_max_row_in_col(self,sh,col)
                    last_row+= 1
                    sh.cell(last_row,col,value=choice) 
                    last_row2 = Friends.get_max_row_in_col(self,sh,col2)  ### to add in friend list of 
                    last_row2+= 1
                    sh.cell(last_row2,col2,value=user)
                    wb.save('Facebook.xlsx')
                    print("You and",choice,'are now friends')

                    Noti = user+' has accepted your friend request'
                    sh2 = wb['Sheet5']
                    l_col = sh2.max_column+1
                    lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                    if choice in lst:
                        index = lst.index(choice)+2
                        last_row = obj3.get_max_row_in_col(sh2,index)+1
                        sh2.cell(last_row,index,value=Noti)
                        wb.save('Facebook.xlsx')
                    
                    Sh = wb['Sheet1']
                    col = Sh.max_column+1
                    name_lst = []
                    for i in range(2,col):
                        val = Sh.cell(1,i).value
                        name_lst.append(val)
                    if user in name_lst:
                        index = name_lst.index(user)
                        col = index+2
                        last_row = Friends.get_max_row_in_col(self,Sh,col)
                        last_row+= 1
                        lst = []
                        for i in range(2,last_row):
                            valu = Sh.cell(i,col).value
                            lst.append(valu)
                        if choice in lst:
                            index = lst.index(choice)
                            row = index+2
                            Sh.cell(row, col, value = NaN)
                            wb.save('Facebook.xlsx')
                else:
                    print('There is no friend request of', choice)
        else:
            print('You are not logged in!')

    def Reject_friend_request(self):
        choice = input('Enter the Email of member whom friend request you want to reject:')
        user = obj1.email
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        col = sh.max_column+1
        name_lst = []
        for i in range(2,col):
            val = sh.cell(1,i).value
            name_lst.append(val)
        if user in name_lst:
            index = name_lst.index(user)
            col = index+2
            last_row = Friends.get_max_row_in_col(self,sh,col)
            last_row+= 1
            lst = []
            for i in range(2,last_row):
                valu = sh.cell(i,col).value
                lst.append(valu)
            if choice in lst:
                sh1 = wb ['Sheet5']
                cc = sh1.max_column
                lst3 =[]
                for i in range(2,cc):
                    valu = sh1.cell(1,i).value
                    lst3.append(valu)
                for j in range(len(lst3)):
                    count = 2
                    if choice == lst3[j]:
                        index = j
                        break
                    else:
                        count+=1
                col3 = index+count
                col3+=1
                last_row = obj3.get_max_row_in_col(sh1,col3)+1
                sh1.cell(last_row, col3,value=user)
                wb.save('Facebook.xlsx')
                index = lst.index(choice)
                row = index+2
                #print(row)
                #print(col)
                sh.cell(row, col, value = NaN)
                wb.save('Facebook.xlsx')
                print('Friend request of', choice,'is rejected')
                self.reject = True

            else:
                print('There is no friend request of',choice)
    def Print_Suggestions(self):
        if obj1.Logged_in == True:
            print('Following are the Suggestions:')
            wb = load_workbook('Facebook.xlsx')
            sh = wb['Sheet']
            last_row = Friends.get_max_row_in_col(self,sh,1)+1
            lst = []
            for row in range(2,last_row):
                valu = sh.cell(row,1).value
                lst.append(valu)

            friend_lst = Friends.Return_friends_lst(self)
            for j in range(len(lst)):
                if lst[j] not in friend_lst:
                    print(lst[j])
    
    def Print_friend_requests(self):
        user = obj1.email
        print('Friend Requests are given below:')
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        L_col = sh.max_column+1
        lst = []
        for i in range(2,L_col):
            valu = sh.cell(1,i).value
            lst.append(valu)
        if user in lst:
            index = lst.index(user)
        col = index+2
        last_row =Friends.get_max_row_in_col(self,sh,col)+1
        for k in range(2,last_row):
            valu = sh.cell(k,col).value
            if valu != None:
                print(valu)

class Post:
    def __init__(self):
        self.obj1 = Facebook()          ##   Composition  ##
        self.obj3 = Friends()


    def Share_post_with_friends(self):
        if obj1.Logged_in == True:
            pp = obj1.name+":"+input('write what you want to share with friends:')
            friends_lst = obj3.Return_friends_lst()
            user = obj1.email
            wb = load_workbook('Facebook.xlsx')
            sh = wb['Sheet4']
            last_col = sh.max_column+1
            l_row = sh.max_row+1
            sh.cell(l_row,1,value=user)
            sh.cell(l_row,2,value=pp)
            wb.save('Facebook.xlsx')
            print('Post has been shared with friends.....')

            Noti = user+' has shared a post'
            sh2 = wb['Sheet5']
            l_col = sh2.max_column+1
            lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
            for i in range(len(lst)):
                for j in range(len(friends_lst)):
                    if friends_lst[j] in lst:
                        if friends_lst[j] != user:
                            index = lst.index(friends_lst[j])+2
                            last_row = obj3.get_max_row_in_col(sh2,index)+1
                            sh2.cell(last_row,index,value=Noti)
                            wb.save('Facebook.xlsx')
                break
        else:
            print('You are not logged in!')


    def Display_post(self):
        user = obj1.email
        friends_lst = obj3.Return_friends_lst()
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet4']
        last_row = sh.max_row+1
        email_lst = obj1.Return_column_values_lst(sh,3,last_row,1)
        post_lst = obj1.Return_column_values_lst(sh,3,last_row,2)
        count = 0
        b = 0
        for j in range(len(post_lst)):
            while True: 
                d = 0
                if email_lst[j] in friends_lst:
                    if post_lst[j] != None:
                        print('Post is given below:')
                        print(post_lst[j])
                        count+=1
                        print('1.Add Comment\n2.See Comments\n3.Next Post\n0.Main menu')
                        choice = input('Make decision:')
                        row = j+3
                        #print(row)
                        if choice == '1':
                            comment = input('Enter your comment:')
                            comment = user+':'+comment
                            col = 3
                            while True:
                                if sh.cell(row,col).value == None:
                                    sh.cell(row,col,value=comment)
                                    wb.save('Facebook.xlsx')
                                    print('Comment has been Added')
                                    if email_lst[j] != user:
                                        Noti = user+' has commented on your post'
                                        sh2 = wb['Sheet5']
                                        l_col = sh2.max_column+1
                                        lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                                        for k in range(len(lst)):
                                            if email_lst[j] in lst:
                                                index = lst.index(email_lst[j])+2
                                                last_row = obj3.get_max_row_in_col(sh2,index)+1
                                                sh2.cell(last_row,index,value=Noti)
                                                wb.save('Facebook.xlsx')
                                                break
                                    break
                                else:
                                    col+=1
                        elif choice == '2':
                            col2 = 3
                            print('Comments are given below:')
                            c =0
                            while True:
                                if sh.cell(row,col2).value != None:
                                    print(sh.cell(row,col2).value)
                                    c+=1
                                    wb.save('Facebook.xlsx')
                                    col2+=1
                                else:
                                    break
                            if c == 0:
                                print('There are no Comments yet!')
                        elif choice == '0':
                            b = 1
                            break       
                        elif choice == '3':
                            d = 1
                        else:
                            print('Invalid input')
                    else:
                        d = 1
                else: 
                    break
                if d == 1:
                    break
            if b == 1:
                break
        if count == 0:
            print('There is no post by you and your friends!')


    def Search_post_by_word(self):
        choice = input('Enter your word to search post:')
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet3']
        last_col =sh.max_column+1
        lst =[]
        for i in range(2,last_col):
            valu = sh.cell(2,i).value
            lst.append(valu)
        for i in range(len(lst)):
            if lst[i] != None:
                if choice in lst[i]:
                    print(lst[i])
            

    

    def See_Notifications(self):
        wb = load_workbook("Facebook.xlsx")
        user = obj1.email
        sh = wb['Sheet5']
        e_col = sh.max_column+1
        header = obj1.Return_Row_values_lst(sh,2,e_col,1)
        if user in header:
            index = header.index(user)+2
        end_row = obj3.get_max_row_in_col(sh,index)+1
        lst = obj1.Return_column_values_lst(sh,2,end_row,index)
        c = len(lst)-1
        if len(lst)!= 0:
            print('Notifications are given below:')
        else:
            print('There is no any notification')
        for n in range(len(lst)):
            print(lst[c])
            c = c-1

class Message:

    def __init__(self):
        self.obj1 = Facebook()          ##   Composition  ##
        self.obj3 = Friends()


    def Sent_message(self):
        wb = load_workbook('Facebook.xlsx')
        sh1 = wb['Sheet']
        user = obj1.email
        receiver = input('Enter the email of member whom you want to send message:')
        Accounts_lst = []
        row = sh1.max_row+1
        for i in range(2,row):
            valu = sh1.cell(i,1).value
            Accounts_lst.append(valu)

        if receiver in Accounts_lst:
            if receiver != user:
                sh = wb['Sheet6']
                message = input('Enter your message:')
                message = user+':'+message
                head = user+' and '+receiver
                head2 = receiver+' and '+user

                sh['A1'] = 'Messages_file'
                last_col = sh.max_column+1
                head_lst = []
                for k in range(2,last_col):
                    valu = sh.cell(1,k).value
                    head_lst.append(valu)
                if head not in head_lst and head2 not in head_lst:
                    sh.cell(1,last_col,value=head)
                    sh.cell(2,last_col,value=message)
                    wb.save('Facebook.xlsx')
                    print('message has been sent')

                    Noti = user+' has sent you a message'
                    sh2 = wb['Sheet5']
                    l_col = sh2.max_column+1
                    lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                    if receiver in lst:
                        index = lst.index(receiver)+2
                        last_row = obj3.get_max_row_in_col(sh2,index)+1
                        sh2.cell(last_row,index,value=Noti)
                        wb.save('Facebook.xlsx')

                elif head in head_lst:
                    index = head_lst.index(head)
                    f_col = index+2
                    sh.cell(2,f_col,value=message)
                    wb.save('Facebook.xlsx')
                    print('message has been sent')

                    Noti = user+' has sent you a message'
                    sh2 = wb['Sheet5']
                    l_col = sh2.max_column+1
                    lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                    if receiver in lst:
                        index = lst.index(receiver)+2
                        last_row = obj3.get_max_row_in_col(sh2,index)+1
                        sh2.cell(last_row,index,value=Noti)
                        wb.save('Facebook.xlsx')
                elif head2 in head_lst:
                    index = head_lst.index(head2)
                    f_col = index+2
                    sh.cell(2,f_col,value=message)
                    wb.save('Facebook.xlsx')
                    print('message has been sent')

                    Noti = user+' has sent you a message'
                    sh2 = wb['Sheet5']
                    l_col = sh2.max_column+1
                    lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                    if receiver in lst:
                        index = lst.index(receiver)+2
                        last_row = obj3.get_max_row_in_col(sh2,index)+1
                        sh2.cell(last_row,index,value=Noti)
                        wb.save('Facebook.xlsx')

            else:
                print("You can't send message to yourself!")
        else:
            print('There is no such member!')

    def Print_message_Suggestions(self):
        user = obj1.email
        print('Following are the Suggestions:')
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet']
        last_row = obj3.get_max_row_in_col(sh,1)+1
        lst = []
        for row in range(2,last_row):
            valu = sh.cell(row,1).value
            lst.append(valu)
        
        for j in range(len(lst)):
            if lst[j] != user:
                print(lst[j])



    def Display_messages(self):
        user = obj1.email
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet6']
        other = input('Enter the email of person whoose messages you want to see:')
        ss = user+' and '+other
        ss2 = other+' and '+user
        last = sh.max_column+1
        lst = []
        for i in range(2,last):
            valu = sh.cell(1,i).value
            lst.append(valu)
        if ss in lst:
            index = lst.index(ss)
            col = index+2
            last_row = obj3.get_max_row_in_col(sh,col)+1
            for r in range(2,last_row):
                valu = sh.cell(r,col).value
                print(valu)
        elif ss2 in lst:
            index = lst.index(ss2)
            col = index+2
            last_row = obj3.get_max_row_in_col(sh,col)+1
            for r in range(2,last_row):
                valu = sh.cell(r,col).value
                print(valu)

        else:
            print('There is no chat in between you and',other)


class Page:
    pass

class Privacy:
    pass


print('....................Welcome to Console based Facebook...................')
obj1= Facebook() 
obj2 = User()
obj3 = Friends()
obj4 = Post()
obj5 = Message()


while True:
    print("1.Create Account\n2.Login\n0.Exit")
    choice = input("Make decision: ")
    if choice == '1':
        obj1.Signup()
    elif choice == '2':
        obj1.Login()
        if obj1.Logged_in == True:
            while True:
                print("1.Add_Bio\n2.Add_Education_details\n3.Add_Work_Experience\n4.Sent_Friend_request\n5.Accept_Friend_request\n6.Reject_Friend_request\n7.Search_Profile_by_name\n8.See_Notifications\n9.Share_privately_with_Friends\n10.Search_Post_with_word\n11.Comment_on_post\n12.Send_message\n13.See_Posts\n14.See_messages\n0.Main menu:")
                choice = input("Make decision: ")
                if choice == '1':
                    obj2.Add_bio()
                    continue
                elif choice == '2':
                    obj2.Add_Education()
                    continue
                elif choice == '3':
                    obj2.Add_Work()
                    continue
                elif choice == '4':
                    obj3.Print_Suggestions()
                    obj3.Send_friend_request()
                    continue
                elif choice == '5':
                    obj3.Print_friend_requests()
                    obj3.Accept_friend_request()
                    continue
                elif choice == '6':
                    obj3.Print_friend_requests()
                    obj3.Reject_friend_request()
                    continue
                elif choice == '7':
                    obj2.Search_member_by_name()
                    continue
                elif choice == '8':
                    obj4.See_Notifications()
                    continue
                elif choice == '9':
                    obj4.Share_post_with_friends()
                    continue
                elif choice == '10':
                    obj4.Search_post_by_word()
                    continue
                elif choice == '11':
                    obj4.Add_Comment()
                    continue
                elif choice == '12':
                    obj5.Print_message_Suggestions()
                    obj5.Sent_message()
                    continue
                elif choice == '13':
                    obj4.Display_post()
                    continue
                elif choice == '14':
                    obj5.Display_messages()
                elif choice == '0':
                    obj1.Logged_in = False
                    break
                else:
                    print('Invalid Input')
    elif choice == '0':
        break
    else:
        print('Invalid input')

        ##################################  The End ####################################
    